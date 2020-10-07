import openpyxl
import struct

from tkinter import *
from tkinter.filedialog import *

popravki = [[0] * 38 for i in range(12)]
from  openpyxl  import  load_workbook

file_out = open('Fab_cable_correction', 'wb')

def write_file():
    for i in range(12):
        for j in range(38):
            out = struct.pack('>h', popravki[i][j])
            file_out.write(out)

def out_console():
    for i in range(12):
        for j in range(38):
            text_pop.insert(i.j,str(popravki[i][j]))

def open_click(event):
    text_pop.delete(0.0,END)
    open_file = askopenfilename()
    wb = load_workbook(open_file)
    sheet = wb['Данные для импорта']
    for i in range(12):
        for j in range(38):
            row = sheet.cell(row=(i + 1), column=(j + 1)).value
            popravki[i][j] = row
            text_pop.insert(0.0, popravki[i][j])
            text_pop.insert(0.0, ' ')
        text_pop.insert(0.0, '\n')

def click_write_file(event):
    write_file()
    text_pop.insert(0.0, '\n')
    text_pop.insert(0.200, 'OK WRITE')
    file_out.close()

def click_exit(event):
    root.destroy()


root = Tk()
root.geometry('1000x600')
root.title('Конвертирование поправок')
bt_open = Button(root, width=8, text='OPEN XLSX')
bt_open.place(x = 30,y = 50)
bt_open.bind('<Button-1>', open_click)

bt_write = Button(root, width=8, text='WRITE_FILE')
bt_write.place(x = 30,y = 100)
bt_write.bind('<Button-1>', click_write_file)

bt_exit = Button(root, width=8, text='EXIT')
bt_exit.place(x = 30,y = 200)
bt_exit.bind('<Button-1>', click_exit)

text_pop = Text(root,height=25,width=100,font='Arial 10',wrap=WORD)
text_pop.place(x = 150,y = 40)
root.mainloop()

